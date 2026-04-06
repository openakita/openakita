"""
表格 API 集成示例
支持 Google Sheets、Excel
"""

from typing import List, Dict, Optional, Any
from datetime import datetime


class GoogleSheetsAPI:
    """Google Sheets API 集成"""
    
    def __init__(self, credentials_file: str = 'credentials.json'):
        self.credentials_file = credentials_file
        self.service = None
    
    def authenticate(self, scopes: List[str] = None):
        """身份认证"""
        from google.oauth2.credentials import Credentials
        from google_auth_oauthlib.flow import InstalledAppFlow
        from google.auth.transport.requests import Request
        from googleapiclient.discovery import build
        import os.path
        
        scopes = scopes or ['https://www.googleapis.com/auth/spreadsheets']
        
        creds = None
        token_file = 'token_sheets.json'
        
        if os.path.exists(token_file):
            creds = Credentials.from_authorized_user_file(token_file, scopes)
        
        if not creds or not creds.valid:
            if creds and creds.expired and creds.refresh_token:
                creds.refresh(Request())
            else:
                flow = InstalledAppFlow.from_client_secrets_file(
                    self.credentials_file, scopes
                )
                creds = flow.run_local_server(port=0)
            
            with open(token_file, 'w') as token:
                token.write(creds.to_json())
        
        self.service = build('sheets', 'v4', credentials=creds)
        print("✓ Google Sheets 认证成功")
    
    def create_spreadsheet(self, title: str) -> Optional[str]:
        """
        创建电子表格
        
        Args:
            title: 表格标题
            
        Returns:
            str: 表格 ID
        """
        if not self.service:
            self.authenticate()
        
        try:
            spreadsheet = {
                'properties': {
                    'title': title
                }
            }
            
            spreadsheet = self.service.spreadsheets().create(body=spreadsheet).execute()
            sheet_id = spreadsheet.get('spreadsheetId')
            print(f"✓ 表格已创建：{sheet_id}")
            print(f"  访问链接：https://docs.google.com/spreadsheets/d/{sheet_id}")
            return sheet_id
            
        except Exception as e:
            print(f"✗ 创建表格失败：{e}")
            return None
    
    def append_rows(self, spreadsheet_id: str, range_name: str, values: List[List[Any]]) -> bool:
        """
        追加行数据
        
        Args:
            spreadsheet_id: 表格 ID
            range_name: 工作表范围（如 "Sheet1!A1"）
            values: 二维数组数据
            
        Returns:
            bool: 操作是否成功
        """
        if not self.service:
            self.authenticate()
        
        try:
            body = {
                'values': values
            }
            
            result = self.service.spreadsheets().values().append(
                spreadsheetId=spreadsheet_id,
                range=range_name,
                valueInputOption='RAW',
                body=body
            ).execute()
            
            print(f"✓ 已追加 {result.get('updates', {}).get('updatedRows', 0)} 行数据")
            return True
            
        except Exception as e:
            print(f"✗ 追加数据失败：{e}")
            return False
    
    def get_values(self, spreadsheet_id: str, range_name: str) -> List[List[Any]]:
        """
        获取单元格数据
        
        Args:
            spreadsheet_id: 表格 ID
            range_name: 工作表范围（如 "Sheet1!A1:D10"）
            
        Returns:
            list: 二维数组数据
        """
        if not self.service:
            self.authenticate()
        
        try:
            result = self.service.spreadsheets().values().get(
                spreadsheetId=spreadsheet_id,
                range=range_name
            ).execute()
            
            values = result.get('values', [])
            print(f"✓ 获取到 {len(values)} 行数据")
            return values
            
        except Exception as e:
            print(f"✗ 获取数据失败：{e}")
            return []
    
    def update_cells(self, spreadsheet_id: str, range_name: str, values: List[List[Any]]) -> bool:
        """
        更新单元格数据
        
        Args:
            spreadsheet_id: 表格 ID
            range_name: 工作表范围
            values: 二维数组数据
            
        Returns:
            bool: 更新是否成功
        """
        if not self.service:
            self.authenticate()
        
        try:
            body = {
                'values': values
            }
            
            result = self.service.spreadsheets().values().update(
                spreadsheetId=spreadsheet_id,
                range=range_name,
                valueInputOption='RAW',
                body=body
            ).execute()
            
            print(f"✓ 已更新 {result.get('updatedCells', 0)} 个单元格")
            return True
            
        except Exception as e:
            print(f"✗ 更新数据失败：{e}")
            return False


class ExcelAPI:
    """Excel 文件操作 API（使用 openpyxl）"""
    
    def __init__(self, filename: str = None):
        from openpyxl import Workbook, load_workbook
        
        self.Workbook = Workbook
        self.load_workbook = load_workbook
        self.filename = filename
        self.wb = None
        self.ws = None
    
    def create_workbook(self, filename: str, sheet_name: str = "Sheet1"):
        """创建新的 Excel 工作簿"""
        self.wb = self.Workbook()
        self.ws = self.wb.active
        self.ws.title = sheet_name
        self.filename = filename
        print(f"✓ 创建工作簿：{filename}")
    
    def open_workbook(self, filename: str):
        """打开现有 Excel 文件"""
        self.wb = self.load_workbook(filename)
        self.ws = self.wb.active
        self.filename = filename
        print(f"✓ 打开工作簿：{filename}")
    
    def write_headers(self, headers: List[str]):
        """写入表头"""
        for col, header in enumerate(headers, 1):
            self.ws.cell(row=1, column=col, value=header)
        print(f"✓ 写入表头：{headers}")
    
    def append_row(self, row_data: List[Any]):
        """追加一行数据"""
        self.ws.append(row_data)
    
    def append_rows(self, rows: List[List[Any]]):
        """追加多行数据"""
        for row in rows:
            self.ws.append(row)
        print(f"✓ 追加 {len(rows)} 行数据")
    
    def save(self, filename: str = None):
        """保存文件"""
        filename = filename or self.filename
        if not filename:
            print("✗ 未指定文件名")
            return
        
        self.wb.save(filename)
        print(f"✓ 文件已保存：{filename}")
    
    def read_all(self) -> List[List[Any]]:
        """读取所有数据"""
        data = []
        for row in self.ws.iter_rows(values_only=True):
            data.append(list(row))
        print(f"✓ 读取到 {len(data)} 行数据")
        return data
    
    def read_range(self, start_row: int, end_row: int, start_col: int, end_col: int) -> List[List[Any]]:
        """读取指定范围的数据"""
        data = []
        for row in self.ws.iter_rows(
            min_row=start_row, max_row=end_row,
            min_col=start_col, max_col=end_col,
            values_only=True
        ):
            data.append(list(row))
        return data


# 使用示例
if __name__ == "__main__":
    # Google Sheets
    sheets = GoogleSheetsAPI()
    sheets.authenticate()
    
    # 创建表格
    spreadsheet_id = sheets.create_spreadsheet("项目进度跟踪")
    
    # 写入表头
    sheets.append_rows(
        spreadsheet_id,
        "Sheet1!A1",
        [["任务", "负责人", "状态", "开始日期", "截止日期", "进度"]]
    )
    
    # 写入数据
    sheets.append_rows(
        spreadsheet_id,
        "Sheet1!A2",
        [
            ["需求分析", "张三", "已完成", "2026-03-01", "2026-03-10", "100%"],
            ["系统设计", "李四", "进行中", "2026-03-11", "2026-03-20", "60%"],
            ["开发实现", "王五", "未开始", "2026-03-21", "2026-04-10", "0%"]
        ]
    )
    
    # Excel
    excel = ExcelAPI()
    excel.create_workbook("项目数据.xlsx", "项目数据")
    excel.write_headers(["姓名", "年龄", "部门", "入职日期"])
    excel.append_rows([
        ["张三", 28, "技术部", datetime(2024, 1, 15)],
        ["李四", 32, "产品部", datetime(2023, 6, 1)],
        ["王五", 25, "市场部", datetime(2025, 3, 10)]
    ])
    excel.save()
