"""
电商 AI 系统 ROI 计算模板生成器
生成 Excel 格式的 ROI 计算模板
"""

import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# 创建 workbook
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "ROI 计算模板"

# 定义样式
title_font = Font(name='Microsoft YaHei', size=18, bold=True, color='FFFFFF')
header_font = Font(name='Microsoft YaHei', size=12, bold=True, color='FFFFFF')
normal_font = Font(name='Microsoft YaHei', size=11)
bold_font = Font(name='Microsoft YaHei', size=11, bold=True)

title_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
input_fill = PatternFill(start_color='FEF9E7', end_color='FEF9E7', fill_type='solid')
result_fill = PatternFill(start_color='52BE80', end_color='52BE80', fill_type='solid')

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
right_align = Alignment(horizontal='right', vertical='center')

# ========== 标题 ==========
ws.merge_cells('A1:E1')
ws['A1'] = '电商 AI 系统投资回报率 (ROI) 计算模板'
ws['A1'].font = title_font
ws['A1'].fill = title_fill
ws['A1'].alignment = center_align

ws['A2'] = f'生成日期：{datetime.now().strftime("%Y-%m-%d")}'
ws['A2'].font = normal_font
ws.merge_cells('B2:E2')
ws['B2'] = '使用说明：黄色单元格为输入项，绿色单元格为计算结果'
ws['B2'].font = normal_font
ws['B2'].alignment = center_align

# ========== 第一部分：人力成本节省计算 ==========
ws['A4'] = '一、人力成本节省计算'
ws['A4'].font = header_font
ws['A4'].fill = header_fill
ws.merge_cells('B4:E4')

headers_1 = ['项目', '当前状态（引入前）', '预期状态（引入后）', '计算公式', '年度节省']
for col, header in enumerate(headers_1, 1):
    cell = ws.cell(row=5, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

data_1 = [
    ('客服团队人数（人）', '5', '2', 'C6*B6'),
    ('客服人均月薪（元）', '6000', '6000', '-'),
    ('客服年度总成本（元）', '=B6*B7*12', '=C6*C7*12', 'B8-C8'),
    ('订单处理人数（人）', '3', '1', 'C9*B9'),
    ('订单人均月薪（元）', '7000', '7000', '-'),
    ('订单年度总成本（元）', '=B9*B10*12', '=C9*C10*12', 'B11-C11'),
    ('数据分析人数（人）', '2', '0.5', 'C12*B12'),
    ('数据人均月薪（元）', '8000', '8000', '-'),
    ('数据年度总成本（元）', '=B12*B13*12', '=C12*C13*12', 'B14-C14'),
    ('培训成本（元/年）', '30000', '10000', 'B15-C15'),
    ('管理成本（元/年）', '50000', '20000', 'B16-C16'),
    ('人力成本总计（元/年）', '=SUM(B8:B16)', '=SUM(C8:C16)', 'B17-C17'),
]

for row_idx, (item, current, expected, formula) in enumerate(data_1, 6):
    ws.cell(row=row_idx, column=1, value=item).font = normal_font
    ws.cell(row=row_idx, column=1).border = thin_border
    ws.cell(row=row_idx, column=1).alignment = left_align
    
    cell_b = ws.cell(row=row_idx, column=2, value=current)
    cell_b.font = normal_font
    if not current.startswith('='):
        cell_b.fill = input_fill
    cell_b.border = thin_border
    cell_b.alignment = right_align
    
    cell_c = ws.cell(row=row_idx, column=3, value=expected)
    cell_c.font = normal_font
    if not expected.startswith('='):
        cell_c.fill = input_fill
    cell_c.border = thin_border
    cell_c.alignment = right_align
    
    ws.cell(row=row_idx, column=4, value=formula).font = normal_font
    ws.cell(row=row_idx, column=4).border = thin_border
    ws.cell(row=row_idx, column=4).alignment = center_align

# 年度节省列
for row_idx in range(6, 18):
    cell_e = ws.cell(row=row_idx, column=5)
    if row_idx == 17:
        cell_e.value = '=B17-C17'
    elif row_idx in [8, 11, 14]:
        cell_e.value = f'=B{row_idx}-C{row_idx}'
    else:
        cell_e.value = '-'
    cell_e.font = bold_font
    cell_e.fill = result_fill
    cell_e.border = thin_border
    cell_e.alignment = right_align

# ========== 第二部分：错误率降低价值 ==========
ws['A18'] = '二、错误率降低价值'
ws['A18'].font = header_font
ws['A18'].fill = header_fill
ws.merge_cells('B18:E18')

headers_2 = ['错误类型', '当前错误率', '预期错误率', '单次错误成本（元）', '年度节省']
for col, header in enumerate(headers_2, 1):
    cell = ws.cell(row=19, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

data_2 = [
    ('订单错误（发错货/漏发）', '3.5%', '0.5%', '150'),
    ('库存错误（超卖/缺货）', '2.0%', '0.3%', '200'),
    ('客服投诉', '5.0%', '1.0%', '100'),
    ('数据错误（报表错误）', '4.0%', '0.5%', '300'),
    ('营销错误（推送错误）', '3.0%', '0.5%', '500'),
]

for row_idx, (item, current, expected, cost) in enumerate(data_2, 20):
    ws.cell(row=row_idx, column=1, value=item).font = normal_font
    ws.cell(row=row_idx, column=1).border = thin_border
    ws.cell(row=row_idx, column=1).alignment = left_align
    
    for col_idx, value in enumerate([current, expected, cost], 2):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.font = normal_font
        cell.fill = input_fill
        cell.border = thin_border
        cell.alignment = right_align

# 年度节省公式
ws.cell(row=20, column=5, value='=(B20-C20)*D20*B27').font = bold_font
ws.cell(row=20, column=5).fill = result_fill
ws.cell(row=20, column=5).border = thin_border
ws.cell(row=20, column=5).alignment = right_align

ws.cell(row=21, column=5, value='=(B21-C21)*D21*B27').font = bold_font
ws.cell(row=21, column=5).fill = result_fill
ws.cell(row=21, column=5).border = thin_border
ws.cell(row=21, column=5).alignment = right_align

ws.cell(row=22, column=5, value='=(B22-C22)*D22*B28').font = bold_font
ws.cell(row=22, column=5).fill = result_fill
ws.cell(row=22, column=5).border = thin_border
ws.cell(row=22, column=5).alignment = right_align

ws.cell(row=23, column=5, value='=(B23-C23)*D23*B29').font = bold_font
ws.cell(row=23, column=5).fill = result_fill
ws.cell(row=23, column=5).border = thin_border
ws.cell(row=23, column=5).alignment = right_align

ws.cell(row=24, column=5, value='=(B24-C24)*D24*B30').font = bold_font
ws.cell(row=24, column=5).fill = result_fill
ws.cell(row=24, column=5).border = thin_border
ws.cell(row=24, column=5).alignment = right_align

ws.cell(row=25, column=5, value='=SUM(E20:E24)').font = bold_font
ws.cell(row=25, column=5).fill = result_fill
ws.cell(row=25, column=5).border = thin_border
ws.cell(row=25, column=5).alignment = right_align

# 业务量输入
ws['A26'] = '业务量输入（用于计算错误成本）'
ws['A26'].font = bold_font
ws.merge_cells('B26:E26')

data_volume = [
    ('年度订单总量（单）', '100000'),
    ('年度咨询总量（次）', '50000'),
    ('年度报表总量（份）', '365'),
    ('年度营销活动总量（次）', '120'),
]

for row_idx, (item, value) in enumerate(data_volume, 27):
    ws.cell(row=row_idx, column=1, value=item).font = normal_font
    ws.cell(row=row_idx, column=1).border = thin_border
    cell_b = ws.cell(row=row_idx, column=2, value=value)
    cell_b.font = normal_font
    cell_b.fill = input_fill
    cell_b.border = thin_border

# ========== 第三部分：效率提升换算 ==========
ws['A32'] = '三、效率提升换算'
ws['A32'].font = header_font
ws['A32'].fill = header_fill
ws.merge_cells('B32:E32')

headers_3 = ['效率指标', '当前水平', '预期水平', '提升比例', '年度价值']
for col, header in enumerate(headers_3, 1):
    cell = ws.cell(row=33, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

data_3 = [
    ('订单处理时效（小时/单）', '4', '0.5'),
    ('客服响应时间（秒）', '47', '8'),
    ('库存周转天数（天）', '45', '30'),
    ('报表制作时间（小时/天）', '3', '0.5'),
    ('营销活动执行时间（小时/次）', '8', '2'),
]

for row_idx, (item, current, expected) in enumerate(data_3, 34):
    ws.cell(row=row_idx, column=1, value=item).font = normal_font
    ws.cell(row=row_idx, column=1).border = thin_border
    ws.cell(row=row_idx, column=1).alignment = left_align
    
    cell_b = ws.cell(row=row_idx, column=2, value=current)
    cell_b.font = normal_font
    cell_b.fill = input_fill
    cell_b.border = thin_border
    cell_b.alignment = right_align
    
    cell_c = ws.cell(row=row_idx, column=3, value=expected)
    cell_c.font = normal_font
    cell_c.fill = input_fill
    cell_c.border = thin_border
    cell_c.alignment = right_align
    
    cell_d = ws.cell(row=row_idx, column=4, value=f'=(B{row_idx}-C{row_idx})/B{row_idx}')
    cell_d.font = bold_font
    cell_d.fill = result_fill
    cell_d.border = thin_border
    cell_d.alignment = right_align
    
    ws.cell(row=row_idx, column=5, value='根据业务情况估算').font = normal_font
    ws.cell(row=row_idx, column=5).border = thin_border
    ws.cell(row=row_idx, column=5).alignment = left_align

# ========== 第四部分：综合 ROI 分析 ==========
ws['A40'] = '四、综合 ROI 分析'
ws['A40'].font = header_font
ws['A40'].fill = header_fill
ws.merge_cells('B40:E40')

data_4 = [
    ('年度人力成本节省（元）', '=E17'),
    ('年度错误成本降低（元）', '=E25'),
    ('效率提升带来的额外收益（元）', '待估算'),
    ('年度总收益（元）', '=SUM(B41:B43)'),
    ('', ''),
    ('AI 系统投入成本（元/年）', ''),
    ('  - 软件许可费', '200000'),
    ('  - 实施费用', '50000'),
    ('  - 培训费用', '10000'),
    ('  - 维护费用', '20000'),
    ('年度总投入（元）', '=SUM(B47:B50)'),
    ('', ''),
    ('投资回报率 ROI', '=(B44-B51)/B51*100'),
    ('投资回收期（月）', '=B51/(B44/12)'),
    ('三年总收益（元）', '=B44*3'),
    ('三年净收益（元）', '=B55-B51*3'),
]

for row_idx, (item, value) in enumerate(data_4, 41):
    if item:
        ws.cell(row=row_idx, column=1, value=item).font = bold_font if 'ROI' in item or '回收期' in item or '净收益' in item else normal_font
        ws.cell(row=row_idx, column=1).border = thin_border
        ws.cell(row=row_idx, column=1).alignment = left_align
    
    if value:
        cell_b = ws.cell(row=row_idx, column=2, value=value)
        cell_b.font = bold_font if 'ROI' in item or '回收期' in item or '净收益' in item else normal_font
        if 'ROI' in item or '回收期' in item or '净收益' in item:
            cell_b.fill = result_fill
        elif not value.startswith('=') and value != '待估算':
            cell_b.fill = input_fill
        cell_b.border = thin_border
        cell_b.alignment = right_align

# ========== 第五部分：敏感性分析 ==========
ws['A58'] = '五、敏感性分析（不同场景下的 ROI）'
ws['A58'].font = header_font
ws['A58'].fill = header_fill
ws.merge_cells('B58:E58')

headers_5 = ['场景', '收益', '投入', 'ROI', '建议']
for col, header in enumerate(headers_5, 1):
    cell = ws.cell(row=59, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border

data_5 = [
    ('乐观场景（收益 +20%）', '=B44*1.2', '=B51', '=(C60-D60)/D60*100', '强烈推荐'),
    ('基准场景（当前估算）', '=B44', '=B51', '=(C61-D61)/D61*100', '推荐'),
    ('保守场景（收益 -20%）', '=B44*0.8', '=B51', '=(C62-D62)/D62*100', '谨慎推荐'),
    ('最差场景（收益 -30%，投入 +20%）', '=B44*0.7', '=B51*1.2', '=(C63-D63)/D63*100', '重新评估'),
]

for row_idx, (scenario, revenue, cost, roi, suggestion) in enumerate(data_5, 60):
    ws.cell(row=row_idx, column=1, value=scenario).font = normal_font
    ws.cell(row=row_idx, column=1).border = thin_border
    ws.cell(row=row_idx, column=1).alignment = left_align
    
    ws.cell(row=row_idx, column=2, value=revenue).font = normal_font
    ws.cell(row=row_idx, column=2).border = thin_border
    ws.cell(row=row_idx, column=2).alignment = right_align
    
    ws.cell(row=row_idx, column=3, value=cost).font = normal_font
    ws.cell(row=row_idx, column=3).border = thin_border
    ws.cell(row=row_idx, column=3).alignment = right_align
    
    cell_roi = ws.cell(row=row_idx, column=4, value=roi)
    cell_roi.font = bold_font
    cell_roi.fill = result_fill
    cell_roi.border = thin_border
    cell_roi.alignment = center_align
    
    cell_sug = ws.cell(row=row_idx, column=5, value=suggestion)
    cell_sug.font = bold_font
    cell_sug.border = thin_border
    cell_sug.alignment = center_align

# 调整列宽
ws.column_dimensions['A'].width = 28
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 18
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 20

# 保存文件
output_path = r'd:\coder\myagent\data\output\电商 AI 系统 ROI 计算模板.xlsx'
wb.save(output_path)

print(f"✅ ROI 计算模板已生成：{output_path}")
