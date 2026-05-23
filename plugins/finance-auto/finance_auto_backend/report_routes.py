"""Report-generation endpoints (M1 W2 Stage 4).

Mounted onto the same APIRouter created by :func:`routes.build_router` -- we
keep the W1 endpoint surface byte-identical and bolt these on via
:func:`register_report_endpoints` so the W1 file stays small.

Four endpoints:

* ``POST /orgs/{org_id}/reports/{kind}/generate``
* ``GET  /orgs/{org_id}/reports``
* ``GET  /orgs/{org_id}/reports/{report_id}``
* ``GET  /orgs/{org_id}/reports/{report_id}/export?format=xlsx``

Decimals + Chinese labels round-trip through JSON cleanly.
"""

from __future__ import annotations

import json
import logging
import tempfile
from pathlib import Path
from typing import TYPE_CHECKING, Any

from fastapi import APIRouter, HTTPException, Query
from fastapi.responses import FileResponse

from .config.yaml_loader import LoadedTemplate, load_template
from .models import (
    ReportCell,
    ReportDetailResponse,
    ReportGenerateRequest,
    ReportInstance,
    ReportListResponse,
)
from .report_generator import (
    GeneratedReport,
    TrialBalanceLine,
    generate_report,
)

if TYPE_CHECKING:
    from .routes import FinanceAutoService

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Template-resolution helpers
# ---------------------------------------------------------------------------


_TEMPLATE_DIR_CANDIDATES = (
    Path(__file__).resolve().parent.parent / "templates" / "reports",
)
"""Where the YAML templates live.  Relative to the backend package; the host
plugin loader resolves the same path because plugin.py copies templates/
alongside the backend code into the runtime tree."""

_KIND_TO_FILENAME = {
    "balance_sheet:small_enterprise": "balance_sheet_small_enterprise.yaml",
    "balance_sheet:general_enterprise": "balance_sheet_general_enterprise.yaml",
    "income_statement:small_enterprise": "income_statement_small_enterprise.yaml",
    "income_statement:general_enterprise": "income_statement_general_enterprise.yaml",
}


def _resolve_template_path(kind: str, standard: str) -> Path:
    key = f"{kind}:{standard}"
    filename = _KIND_TO_FILENAME.get(key)
    if filename is None:
        raise HTTPException(
            status_code=400,
            detail=f"unsupported (kind, standard) combination: {key!r}",
        )
    for root in _TEMPLATE_DIR_CANDIDATES:
        candidate = root / filename
        if candidate.exists():
            return candidate
    raise HTTPException(
        status_code=500, detail=f"template file not found: {filename}"
    )


def _standard_for_org(org_standard: str, override: str | None) -> str:
    if override:
        return override
    if org_standard == "small":
        return "small_enterprise"
    return "general_enterprise"


# ---------------------------------------------------------------------------
# DB helpers (leveraging the W1 service for connection access)
# ---------------------------------------------------------------------------


async def _load_balance_lines(
    service: FinanceAutoService,
    *,
    org_id: str,
    period_id: str,
    source_import_id: str | None,
) -> tuple[list[TrialBalanceLine], str]:
    if source_import_id is None:
        async with service.db.conn.execute(
            "SELECT id FROM trial_balance_imports "
            "WHERE org_id=? AND period_id=? AND status='ok' "
            "ORDER BY uploaded_at DESC LIMIT 1",
            (org_id, period_id),
        ) as cur:
            row = await cur.fetchone()
        if row is None:
            raise HTTPException(
                status_code=404,
                detail=(
                    f"no successful balance-table import for org={org_id} "
                    f"period={period_id}"
                ),
            )
        source_import_id = row[0]

    rows_typed = await service.list_all_rows(
        org_id=org_id, import_id=source_import_id
    )
    lines = [
        TrialBalanceLine(
            id=r.id,
            full_code=r.full_code,
            parent_code=r.parent_code,
            child_code=r.child_code,
            account_name=r.account_name,
            opening_debit=r.opening_debit,
            opening_credit=r.opening_credit,
            period_debit=r.period_debit,
            period_credit=r.period_credit,
            closing_debit=r.closing_debit,
            closing_credit=r.closing_credit,
        )
        for r in rows_typed
    ]
    return lines, source_import_id


async def _persist_report(
    service: FinanceAutoService,
    *,
    template: LoadedTemplate,
    generated: GeneratedReport,
) -> None:
    inst = generated.instance
    await service.db.conn.execute(
        "INSERT INTO reports(id, org_id, period_id, sheet_kind, "
        "accounting_standard, template_id, template_version, status, "
        "cell_count, warnings_json, source_import_id, backend_used, "
        "output_path, generated_at) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        (
            inst.id,
            inst.org_id,
            inst.period_id,
            inst.sheet_kind,
            inst.accounting_standard,
            inst.template_id,
            inst.template_version,
            inst.status,
            inst.cell_count,
            json.dumps(inst.warnings, ensure_ascii=False),
            inst.source_import_id,
            inst.backend_used,
            inst.output_path,
            inst.generated_at,
        ),
    )
    await service.db.conn.executemany(
        "INSERT INTO report_cells(id, report_id, reference_code, target_line_no, "
        "target_label, indent_level, data_source, code, value, sign, is_total, "
        "is_tbd, formula, notes, source_rows) "
        "VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)",
        [
            (
                c.id,
                c.report_id,
                c.reference_code,
                c.target_line_no,
                c.target_label,
                c.indent_level,
                c.data_source,
                c.code,
                c.value,
                c.sign,
                int(c.is_total),
                int(c.is_tbd),
                c.formula,
                c.notes,
                json.dumps(c.source_rows, ensure_ascii=False),
            )
            for c in generated.cells
        ],
    )
    await service.db.conn.commit()


def _row_to_report(row: Any) -> ReportInstance:
    warnings = json.loads(row["warnings_json"] or "[]")
    return ReportInstance(
        id=row["id"],
        org_id=row["org_id"],
        period_id=row["period_id"],
        sheet_kind=row["sheet_kind"],
        accounting_standard=row["accounting_standard"],
        template_id=row["template_id"],
        template_version=row["template_version"] or 1,
        status=row["status"] or "ok",
        cell_count=row["cell_count"] or 0,
        warnings=warnings,
        source_import_id=row["source_import_id"],
        backend_used=row["backend_used"],
        output_path=row["output_path"],
        generated_at=row["generated_at"],
    )


def _row_to_cell(row: Any) -> ReportCell:
    sources = json.loads(row["source_rows"] or "[]")
    return ReportCell(
        id=row["id"],
        report_id=row["report_id"],
        reference_code=row["reference_code"],
        target_line_no=row["target_line_no"],
        target_label=row["target_label"],
        indent_level=row["indent_level"],
        data_source=row["data_source"],
        code=row["code"],
        value=row["value"] or 0.0,
        sign=row["sign"] or 1,
        is_total=bool(row["is_total"]),
        is_tbd=bool(row["is_tbd"]),
        formula=row["formula"],
        notes=row["notes"],
        source_rows=sources,
    )


# ---------------------------------------------------------------------------
# Excel export (programmatic openpyxl writer; the YAML xltpl_file references
# remain forward-looking until the design team ships hand-laid templates).
# ---------------------------------------------------------------------------


def _build_workbook(template: LoadedTemplate, instance: ReportInstance,
                    cells: list[ReportCell]) -> Path:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill

    fd, name = tempfile.mkstemp(suffix=".xlsx", prefix="finauto_report_")
    import os as _os
    _os.close(fd)
    out = Path(name)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = template.sheet_kind[:31]

    title = (
        f"{template.name} - {instance.period_id} "
        f"({instance.accounting_standard})"
    )
    ws.merge_cells("A1:D1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")

    headers = ["项目", "代码", "金额", "备注"]
    for i, h in enumerate(headers, start=1):
        c = ws.cell(row=2, column=i, value=h)
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", fgColor="DDEBF7")

    cells_sorted = sorted(cells, key=lambda c: (c.target_line_no, c.reference_code))
    section_font = Font(bold=True, color="305496")
    total_font = Font(bold=True)
    total_fill = PatternFill("solid", fgColor="FFF2CC")
    tbd_fill = PatternFill("solid", fgColor="F8CBAD")

    row_idx = 3
    for cell in cells_sorted:
        indent = "  " * (cell.indent_level or 0)
        label = f"{indent}{cell.target_label}"
        ws.cell(row=row_idx, column=1, value=label)
        ws.cell(row=row_idx, column=2, value=cell.code or cell.reference_code)
        ws.cell(row=row_idx, column=3, value=float(cell.value))
        ws.cell(row=row_idx, column=3).number_format = "#,##0.00"
        notes = cell.notes or ""
        if cell.is_tbd:
            notes = f"[TBD] {notes}".strip()
        ws.cell(row=row_idx, column=4, value=notes)
        if cell.data_source == "section":
            ws.cell(row=row_idx, column=1).font = section_font
        if cell.is_total:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).font = total_font
                ws.cell(row=row_idx, column=col).fill = total_fill
        if cell.is_tbd:
            for col in range(1, 5):
                ws.cell(row=row_idx, column=col).fill = tbd_fill
        row_idx += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 36
    wb.save(str(out))
    wb.close()
    return out


# ---------------------------------------------------------------------------
# Endpoint registration
# ---------------------------------------------------------------------------


def register_report_endpoints(
    router: APIRouter, service: FinanceAutoService
) -> None:
    @router.post(
        "/orgs/{org_id}/reports/{kind}/generate",
        status_code=201,
        summary="按 YAML 模板生成主表（小企业 / 企业准则资产负债表 + 利润表）",
    )
    async def generate(
        org_id: str, kind: str, payload: ReportGenerateRequest
    ) -> ReportDetailResponse:
        if kind not in {"balance_sheet", "income_statement"}:
            raise HTTPException(
                status_code=400,
                detail=f"unsupported report kind: {kind!r}; allowed: "
                "balance_sheet | income_statement",
            )
        org = await service.get_org(org_id)
        standard = _standard_for_org(org.standard, payload.accounting_standard)
        template_path = _resolve_template_path(kind, standard)
        template = load_template(template_path)

        balance_lines, source_id = await _load_balance_lines(
            service,
            org_id=org_id,
            period_id=payload.period_id,
            source_import_id=payload.source_import_id,
        )
        generated = generate_report(
            template=template,
            org_id=org_id,
            period_id=payload.period_id,
            accounting_standard=standard,
            balance_lines=balance_lines,
            source_import_id=source_id,
        )
        await _persist_report(service, template=template, generated=generated)

        return ReportDetailResponse(
            report=generated.instance, cells=generated.cells
        )

    @router.get(
        "/orgs/{org_id}/reports",
        summary="列出某账套已生成的报表实例",
    )
    async def list_reports(org_id: str) -> ReportListResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? ORDER BY generated_at DESC",
            (org_id,),
        ) as cur:
            rows = await cur.fetchall()
        items = [_row_to_report(r) for r in rows]
        return ReportListResponse(reports=items, total=len(items))

    @router.get(
        "/orgs/{org_id}/reports/{report_id}",
        summary="读取一份报表（含全部 ReportCell）",
    )
    async def get_report(org_id: str, report_id: str) -> ReportDetailResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? AND id=?",
            (org_id, report_id),
        ) as cur:
            row = await cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="report not found")
        instance = _row_to_report(row)
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE report_id=? "
            "ORDER BY target_line_no ASC",
            (report_id,),
        ) as cur:
            cell_rows = await cur.fetchall()
        cells = [_row_to_cell(r) for r in cell_rows]
        return ReportDetailResponse(report=instance, cells=cells)

    @router.get(
        "/orgs/{org_id}/reports/{report_id}/export",
        summary="导出报表为 Excel (.xlsx)",
        response_class=FileResponse,
    )
    async def export_report(
        org_id: str,
        report_id: str,
        format: str = Query(default="xlsx", pattern="^(xlsx)$"),
    ) -> FileResponse:
        await service.get_org(org_id)
        async with service.db.conn.execute(
            "SELECT * FROM reports WHERE org_id=? AND id=?",
            (org_id, report_id),
        ) as cur:
            row = await cur.fetchone()
        if row is None:
            raise HTTPException(status_code=404, detail="report not found")
        instance = _row_to_report(row)
        async with service.db.conn.execute(
            "SELECT * FROM report_cells WHERE report_id=? "
            "ORDER BY target_line_no ASC",
            (report_id,),
        ) as cur:
            cell_rows = await cur.fetchall()
        cells = [_row_to_cell(r) for r in cell_rows]

        template_path = _resolve_template_path(
            instance.sheet_kind, instance.accounting_standard
        )
        template = load_template(template_path)
        out_path = _build_workbook(template, instance, cells)

        await service.db.conn.execute(
            "UPDATE reports SET output_path=?, backend_used=? WHERE id=?",
            (str(out_path), "openpyxl", report_id),
        )
        await service.db.conn.commit()

        filename = (
            f"{instance.template_id}_{instance.period_id}_"
            f"{instance.id[-8:]}.xlsx"
        )
        return FileResponse(
            str(out_path),
            media_type=(
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            ),
            filename=filename,
        )
